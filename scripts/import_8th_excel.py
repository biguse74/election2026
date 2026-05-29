#!/usr/bin/env python3
"""선관위 8회 사전투표현황 엑셀 → baseline_8th.json by_sido_hourly 병합.

입력:
    C:\\Users\\bigus\\Downloads\\사전투표현황[제8회][지방선거][1일차].xlsx
    C:\\Users\\bigus\\Downloads\\사전투표현황[제8회][지방선거][2일차].xlsx

엑셀 구조:
    r5: 시도명, 선거인수, 사전투표자수(%)
    r6: (병합) 7시, 8시, …, 18시(1일차) 또는 7~20시(2일차)
    r7~8: '합계' voted / turnout
    r9~r10, r11~r12, … 시도별 voted / turnout 2행 묶음

8회 vs 9회 시도명:
    강원도 → 강원특별자치도, 전라북도 → 전북특별자치도

양일 누적 통일 (위키·9회 페이지와 일관):
    day1[h] = 1일차 h시 누적 (엑셀 그대로)
    day2[h] = 시도 1일차 최종(18시) + 엑셀 2일차 h시 누적

저장 위치:
    data/early_voting/baseline_8th.json 의 by_sido_hourly 필드.
    파싱 검증용으로 national_final도 엑셀 기준값으로 갱신.
"""

from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DOWNLOADS = Path(r"C:\Users\bigus\Downloads")
SRC_D1 = DOWNLOADS / "사전투표현황[제8회][지방선거][1일차].xlsx"
SRC_D2 = DOWNLOADS / "사전투표현황[제8회][지방선거][2일차].xlsx"
OUT_PATH = ROOT / "data" / "early_voting" / "baseline_8th.json"

SIDO_ALIAS = {
    "강원도": "강원특별자치도",
    "전라북도": "전북특별자치도",
}


def fix_and_load(src: Path):
    """xlsx phantom drawing rels 제거 후 openpyxl 로딩."""
    tmp = src.with_name(src.stem + "__fixed.xlsx")
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "xl/worksheets/_rels/Sheet1.xml.rels":
                data = (
                    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>'
                )
            zout.writestr(item, data)
    wb = openpyxl.load_workbook(tmp, data_only=True)
    tmp.unlink()
    return wb


def parse_workbook(path: Path) -> tuple[dict, dict, list[int]]:
    """엑셀 1개 → (national, sido_map, hour_list).

    national:  {hour: turnout%}
    sido_map:  {sdName: {hour: turnout%}}
    hour_list: 컬럼 순서대로 시각 정수 리스트
    """
    wb = fix_and_load(path)
    ws = wb.active

    # r6: 시간 헤더 — 3번째 열부터
    hour_cells = [ws.cell(6, c).value for c in range(3, ws.max_column + 1)]
    hours = []
    for v in hour_cells:
        if isinstance(v, str):
            m = re.match(r"\s*(\d+)\s*시", v)
            if m:
                hours.append(int(m.group(1)))

    def parse_pct(s):
        if s is None:
            return None
        s = str(s).strip().rstrip("%")
        try:
            return round(float(s), 2)
        except ValueError:
            return None

    # r7~8 합계, r9~ 시도별
    national = {}
    sido_map = {}

    for r_voted in range(7, ws.max_row + 1, 2):
        name = ws.cell(r_voted, 1).value
        if name is None:
            continue
        name = str(name).strip()
        r_pct = r_voted + 1
        pcts = {}
        for i, h in enumerate(hours):
            v = parse_pct(ws.cell(r_pct, 3 + i).value)
            if v is not None:
                pcts[h] = v
        if not pcts:
            continue
        if name == "합계":
            national = pcts
        else:
            sd = SIDO_ALIAS.get(name, name)
            sido_map[sd] = pcts

    return national, sido_map, hours


def main():
    print("=" * 60)
    print("8회 지선 사전투표 — 선관위 엑셀 → baseline_8th.json")
    print("=" * 60)

    print(f"\n[1일차] {SRC_D1.name}")
    nat1, sido1, hours1 = parse_workbook(SRC_D1)
    print(f"  · 시간대: {hours1}")
    print(f"  · 합계 row: {nat1}")
    print(f"  · 시도 수: {len(sido1)}")
    sample_sido = next(iter(sido1))
    print(f"  · 샘플({sample_sido}): {sido1[sample_sido]}")

    print(f"\n[2일차] {SRC_D2.name}")
    nat2, sido2, hours2 = parse_workbook(SRC_D2)
    print(f"  · 시간대: {hours2}")
    print(f"  · 합계 row: {nat2}")
    print(f"  · 시도 수: {len(sido2)}")
    print(f"  · 샘플({sample_sido}): {sido2[sample_sido]}")

    # 양일 누적 변환:
    #   day1 누적[h] = 엑셀 1일차 누적[h]
    #   day2 누적[h] = 시도 1일차 최종(18시) + 엑셀 2일차 누적[h]
    # national도 동일 규칙.
    nat_day1_final = nat1.get(18)
    print(f"\n[양일 누적 변환] 전국 1일차 18시 final = {nat_day1_final}%")

    by_sido_hourly = {"day1": {}, "day2": {}}
    sido_day1_finals = {}
    for sd, hmap in sido1.items():
        by_sido_hourly["day1"][sd] = {str(h): v for h, v in sorted(hmap.items())}
        sido_day1_finals[sd] = hmap.get(18)

    for sd, hmap in sido2.items():
        sd_d1 = sido_day1_finals.get(sd)
        if sd_d1 is None:
            print(f"  ! {sd}: 1일차 final 없음 — skip", file=sys.stderr)
            continue
        cum = {str(h): round(sd_d1 + v, 2) for h, v in sorted(hmap.items())}
        by_sido_hourly["day2"][sd] = cum

    # national hourly 갱신: 동일 규칙
    nat_hourly_day1 = [{"hour": h, "cum": nat1[h]} for h in sorted(nat1)]
    nat_hourly_day2 = [{"hour": h, "cum": round(nat_day1_final + v, 2)} for h, v in sorted(nat2.items())]

    # 양일 최종 = day2 마지막 시점 (보통 20시)
    final_cum = nat_hourly_day2[-1]["cum"] if nat_hourly_day2 else None
    print(f"  · 양일 최종 누적 (전국): {final_cum}%")

    # baseline 병합
    base = json.loads(OUT_PATH.read_text(encoding="utf-8"))

    # by_sido_hourly 추가
    base["by_sido_hourly"] = by_sido_hourly

    # hourly_national 갱신 (엑셀 기준으로 정확)
    base["hourly_national"] = {"day1": nat_hourly_day1, "day2": nat_hourly_day2}

    # national_final 갱신
    if final_cum is not None:
        base["national_final"] = final_cum

    # source / note 갱신
    base["source"] = "중앙선거관리위원회 선거통계시스템 사전투표현황 엑셀 (제8회 지방선거 1·2일차)"
    base["note"] = (
        "선관위 공식 엑셀(2026-05-29 수신) 기준. day1=1일차 누적, "
        "day2=양일 누적(1일차 final + 2일차 시간대). "
        "by_sido_hourly에 17 시도×시간대 사전투표율(%) 수록. "
        "시도 명칭은 9회 표준(강원특별자치도·전북특별자치도)으로 보정."
    )

    OUT_PATH.write_text(json.dumps(base, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n저장 → {OUT_PATH.relative_to(ROOT)}")
    print(f"  · day1 시도 슬롯: {sum(len(v) for v in by_sido_hourly['day1'].values())}")
    print(f"  · day2 시도 슬롯: {sum(len(v) for v in by_sido_hourly['day2'].values())}")

    # 검증: 전남·서울·대구 day1 09시 / day2 12시 출력
    print("\n[검증] 주요 시도 누적률")
    for sd in ["서울특별시", "전라남도", "대구광역시", "강원특별자치도", "전북특별자치도"]:
        d1_09 = by_sido_hourly["day1"].get(sd, {}).get("9")
        d2_18 = by_sido_hourly["day2"].get(sd, {}).get("18")
        d2_last_key = sorted(by_sido_hourly["day2"].get(sd, {}).keys(), key=int)[-1] if by_sido_hourly["day2"].get(sd) else None
        d2_last = by_sido_hourly["day2"].get(sd, {}).get(d2_last_key) if d2_last_key else None
        print(f"  {sd}: day1 9시 {d1_09}% / day2 18시 {d2_18}% / day2 최종({d2_last_key}시) {d2_last}%")


if __name__ == "__main__":
    main()
